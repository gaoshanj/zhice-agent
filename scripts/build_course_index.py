"""构建微软官方培训课程知识库（course_docs 集合）

用法:
    python scripts/build_course_index.py                      # 使用默认列表
    python scripts/build_course_index.py AB-620T00 GH-200T00   # 指定课程编号
    python scripts/build_course_index.py --file scripts/course_list.txt
    python scripts/build_course_index.py --xlsx data/course_catalog.xlsx   # 课程表（字段优先）

课程表 xlsx 字段优先级最高；Learn API 仅补充表格没有的受众/大纲。
每次运行清空并重建 course_docs 集合。需要先配置 .env 中的 Azure OpenAI 凭据。
"""
from __future__ import annotations

import asyncio
import argparse
import sys
from pathlib import Path

# 确保项目根目录在 sys.path（脚本位于 scripts/ 下）
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.llm.learn_search import (
    build_course_knowledge_list,
    build_course_from_rows,
    format_course_knowledge,
)
from src.rag.vector_store import clear_collection, add_chunks
from src.utils.config import settings
from src.utils.logger import logger


# Course Map (FY27 Q1) 默认课程列表 —— 用户可从图片补充完整列表到 --file
DEFAULT_COURSES = [
    "AB-620T00",   # Copilot Studio 集成 AI Agent 解决方案（12 天）
    "GH-200T00",    # GitHub Actions 自动化（1 天）
    "DP-604T00",    # Microsoft Fabric 数据科学与 ML（AI）
    "MB-310T00",    # Dynamics 365 Finance
    "MS-4014",      # 创建智能 agent（Course Map）
    "AI-3016",      # 构建 AI Agent（Course Map）
    "PL-7008",      # Copilot Studio（可能为演进编号，找不到则跳过）
]

# 课程表默认路径（仓库内，随代码部署到 Azure）
DEFAULT_XLSX = str(ROOT / "data" / "course_catalog.xlsx")


def _read_course_list(path: str) -> list[str]:
    p = Path(path)
    if not p.exists():
        logger.warning(f"课程列表文件不存在: {path}")
        return []
    nums: list[str] = []
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#"):
            nums.append(line)
    return nums


def _read_course_xlsx(path: str) -> list[dict]:
    """读取课程表 xlsx，返回规范化行列表。

    期望表头（不区分大小写）：Course Number, Title, Duration,
    Detail Page Url, Solution Area, Credential, State 等。
    跳过课程编号为空的行。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("未安装 openpyxl，无法读取课程表：pip install openpyxl")
        return []

    p = Path(path)
    if not p.exists():
        logger.warning(f"课程表文件不存在: {path}")
        return []

    wb = load_workbook(p, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    key_map = {
        "course number": "course_number",
        "title": "title",
        "duration": "duration",
        "detail page url": "detail_page_url",
        "detail page type": "detail_page_type",
        "solution area": "solution_area",
        "credential": "credential",
        "state": "state",
        "pub type": "pub_type",
        "scope": "scope",
        "release date": "release_date",
    }
    norm_header = [key_map.get(h.lower(), h.lower().replace(" ", "_")) for h in header]

    out: list[dict] = []
    for r in rows[1:]:
        row = {norm_header[i]: (r[i] if i < len(r) else None) for i in range(len(norm_header))}
        num = row.get("course_number")
        if num is None or not str(num).strip():
            continue
        clean = {}
        for k, v in row.items():
            if v is None:
                clean[k] = ""
            elif isinstance(v, float) and v.is_integer():
                clean[k] = str(int(v))
            else:
                clean[k] = str(v).strip()
        out.append(clean)
    logger.info(f"课程表解析: {len(out)} 门有效课程（共 {len(rows) - 1} 行）")
    return out


def _write_chunks(parsed_list: list[dict]) -> int:
    """将解析后的课程列表格式化为 chunk 并写入 course_docs（清空后重建，幂等）"""
    if not parsed_list:
        logger.error("没有任何课程构建成功")
        return 0
    chunks: list[dict] = []
    for parsed in parsed_list:
        content = format_course_knowledge(parsed)
        if not content:
            continue
        chunks.append({
            "chunk_id": f"course_{parsed['course_number']}",
            "content": content,
            "metadata": {
                "source": "course",
                "course_number": parsed["course_number"],
                "title": parsed["title"],
                "products": ",".join(parsed.get("products", [])),
                "roles": ",".join(parsed.get("roles", [])),
                "levels": ",".join(parsed.get("levels", [])),
                "duration_days": parsed.get("duration_days", 0),
                "url": parsed.get("url", ""),
                "solution_area": parsed.get("solution_area", ""),
                "credential": parsed.get("credential", ""),
                "state": parsed.get("state", ""),
            },
        })

    # 清空旧集合后重建（保证幂等）
    clear_collection(settings.chroma_collection_course)
    n = add_chunks(chunks, collection_name=settings.chroma_collection_course)
    logger.info(f"课程知识库构建完成: {n} 门课程写入 {settings.chroma_collection_course}")
    return n


async def run_build(course_numbers: list[str], locale: str = "en-us") -> int:
    """构建课程知识库核心逻辑（按课程编号），返回写入的课程数"""
    if not course_numbers:
        logger.error("没有要构建的课程编号")
        return 0
    logger.info(f"开始构建课程知识库，共 {len(course_numbers)} 门课程")
    parsed_list = await build_course_knowledge_list(course_numbers, locale=locale)
    return _write_chunks(parsed_list)


async def run_build_from_xlsx(xlsx_path: str, locale: str = "en-us") -> int:
    """从课程表 xlsx 构建课程知识库（表格字段优先级最高），返回写入的课程数"""
    rows = _read_course_xlsx(xlsx_path)
    if not rows:
        logger.error("课程表解析为空，终止")
        return 0
    logger.info(f"开始从课程表构建，共 {len(rows)} 行")
    parsed_list = await build_course_from_rows(rows, locale=locale)
    return _write_chunks(parsed_list)


def main() -> None:
    parser = argparse.ArgumentParser(description="构建微软培训课程知识库 course_docs")
    parser.add_argument("courses", nargs="*", help="课程编号（可选）")
    parser.add_argument("--file", help="课程编号文件（每行一个）")
    parser.add_argument("--xlsx", help="课程表 xlsx 路径（表格字段优先级最高）")
    parser.add_argument("--locale", default="en-us")
    args = parser.parse_args()

    if args.xlsx:
        n = asyncio.run(run_build_from_xlsx(args.xlsx, locale=args.locale))
    elif args.courses:
        n = asyncio.run(run_build(args.courses, locale=args.locale))
    elif args.file:
        course_numbers = _read_course_list(args.file)
        if not course_numbers:
            logger.error("课程列表文件为空，终止")
            sys.exit(1)
        n = asyncio.run(run_build(course_numbers, locale=args.locale))
    else:
        n = asyncio.run(run_build(DEFAULT_COURSES, locale=args.locale))

    if n == 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
